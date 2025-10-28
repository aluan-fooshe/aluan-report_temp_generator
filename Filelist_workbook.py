# ----------------------------------------
# Name    : Audrey
# Note    : This is for creating a xlsx file for listing a bunch of files in a single folder by last write time.
#
# Date Created : July 15, 2025 @8:50PM
# ----------------------------------------

# public library imports
import sys
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

dictionary_file1 = 'name_of_file.txt'
dictionary_file2 = 'last_writetime.txt'

class Excel_Filelist:
    """
    A class to handle file operations and Excel spreadsheet management.
    """
    def __init__(self, worksheet=None, dictionary_file1='name_of_file.txt', dictionary_file2='last_writetime.txt'):
        self.ws = worksheet
        self.dictionary_file1 = dictionary_file1
        self.dictionary_file2 = dictionary_file2
        self.dictionaries = {}

    def import_dictionary(self, filename):
        dictionary = {}
        item0 = []

        try:
            f = open(filename, 'r')
            list = f.readlines()
            for item in list:
                item = item.strip('\n')
                item0.append(item)
            #print(item0)

            index = 0
            for item in item0:
                dictionary.update({index: item})
                index += 1

        except:
            print('list_of_files.txt file does not exist')
        return dictionary

    def set_column_width_pixels(self, col_letter, width=8.43):
        """
        Set column width using pixel measurement.

        Parameters:
        col_letter: column letter (e.g., 'A', 'B', 'C')
            col_letter = 'A'
        width = 8.43 # default column length in xlsx sheets?
        """
        ws.column_dimensions[col_letter].width = 2 * width
        return ws.column_dimensions[col_letter].width


    def print_dictionary(self, dictionary):
        for item in dictionary.items():
            print(f"\t{item}")

def add_to_spreadsheet(dictionary, letter):
    return_str = ""
    i = 3
    for key, value in dictionary.items():
        name_cell = f"{letter}{i}"
        ws[name_cell] = f"{value}"
        return_str = return_str + f"{name_cell}\t{key}:{value}\n"
        i += 1
    return return_str

if __name__ == '__main__':

    print(sys.executable)
    print(sys.path)
    #
    # filelist_wb = Workbook()
    # ws = filelist_wb.active
    # ws.title = "List of Files"
    #
    # ws['B1'] = "Filelist of folder"
    # ws['C1'] = "2025-07-16 11:04AM"
    # ws['A2'] = "Image"
    # ws['B2'] = "Name"
    # ws['C2'] = "LastWriteTime"
    #
    # excel_fl = Excel_Filelist('filelist.xlsx', 'last_writetime.txt', 'last_writetime.txt')
    #
    # width = 10
    # excel_fl.set_column_width_pixels('A', width)
    # excel_fl.set_column_width_pixels('B', width*2)
    # excel_fl.set_column_width_pixels('C', width*1.5)
    #
    # # Add image to my filelist spreadsheet
    # image_path = r"C:\Users\Audrey\OneDrive\Pictures\screenshot-collages\2025-02-07 144957 switch_before_go.png"
    # cell_address = "A3"
    #
    # # Load with Pillow first to get pixel dimensions
    # pil_img = PILImage.open(image_path)
    # width_px, height_px = pil_img.size
    # print(f"Original Pixel dimensions: {width_px} x {height_px}")
    #
    # # Convert Excel column width → pixels (approximation)
    # scale = 100
    # # Scale height proportionally, height/width
    # proportional_factor = height_px / width_px
    # target_height_px = int(proportional_factor * scale)
    # print(f"New Pixel dimensions: {scale} x {target_height_px}")
    #
    # # (width, height) makes image proportional to uniform width for xlsx sheet.
    # pil_img = pil_img.resize((scale, target_height_px))
    # saved_image_path = r"C:\Users\Audrey\OneDrive\Pictures\screenshot-resized100\2025-02-07 144957 switch_before_go.png"
    # pil_img.save(saved_image_path)
    #
    # # Load into openpyxl and anchor
    # img = Image(saved_image_path)
    # # Shows the image in image viewer
    # pil_img.show()
    #
    # img.anchor = cell_address
    # ws.row_dimensions[3].height = target_height_px
    # ws.add_image(img)
    #
    #
    # i = 3
    # dictionary1 = excel_fl.import_dictionary(dictionary_file1)
    # dictionary2 = excel_fl.import_dictionary(dictionary_file2)
    # key1, value1 = next(iter(dictionary1.items()))
    # key2, value2 = next(iter(dictionary2.items()))
    #
    # name_cell = f"B{i}"
    # lastwritetime_cell = f"C{i}"
    # ws[name_cell] = f"{value1}"
    # ws[lastwritetime_cell] = f"{value2}"
    # print(f"{name_cell}\t{key1}:{value1}")
    # print(f"{lastwritetime_cell}\t{key2}:{value2}")
    #
    # excel_fl.print_dictionary(dictionary1)
    # print("\n")
    # excel_fl.print_dictionary(dictionary2)
    #
    # print("--------------------\n")
    #
    # add_to_spreadsheet(dictionary1, "B")
    # add_to_spreadsheet(dictionary2, "C")
    #
    # filelist_wb.save('filelist.xlsx')
    # print(f"\n{ws.title} spreadsheet is saved!")